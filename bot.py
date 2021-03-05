import random
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import re
from tkinter import *
from tkinter import ttk
import os
from datetime import date
from functools import partial
#imports from test.py
from test import get_yesterday_date, status_attdendance2
import operator
import pandas as pd



class appbot:



    def __init__(self, master):
        self.master = master

        self.row = 0
        self.column = 0
        self.user_entered = ''
        self.item_text = ''

        self.wb = Workbook()
        self.txt = Text(master)
        self.entry = ttk.Entry(master)

        self.today = date.today()




    def pressed_enter(self,event=None):                                     # send message after hitting enter key            
            self.get_text()

    def send_greeting(self):                                                # a message about the bot
        self.txt.insert(END, "BOT : Hello I am Yash's bot\nI am programmed to answer\n20+ Q&A, Excited to help you!!!!\n")

    def send(self):                                                         # a suggestion after send_greeting msg
            
            cwd = os.getcwd()
            join = os.path.join(cwd, 'Files')
            dirs = os.listdir(join)
            # print(dirs)

            self.txt.insert(END, 'BOT : Select one from below\n', 'start0')
            self.txt.tag_config("start0", background="white", foreground="red")
            for i in dirs:
                # self.item_text = i

                button = Button(self.txt, text=i, padx=2, pady=2,
                        cursor="left_ptr",
                        bd=1, highlightthickness=4,
                        command = partial(self.get_item_text, i))
                
                self.txt.window_create("end", window=button)

    def get_item_text(self, i):                                             # get folder name, selected by user
        self.item_text = i
        self.txt.insert(END, "\n" + f"BOT : You can proceed with {i}")

        self.suggestions(['totall present', 'totall absent'])


    def suggestions(self, lst):                                             # suggestions tags
       
        self.txt.insert(END,'\n'+"BOT : suggestions->", 'start')
        self.txt.tag_config("start", background="white", foreground="gray34")
        
        for i in lst: 
                button = Button(self.txt, text=i, padx=2, pady=2,
                        cursor="left_ptr",
                        bd=1, highlightthickness=0,
                        command = partial(self.greetings , i))
                
                self.txt.window_create("end", window=button)

    def get_text(self):                                                      # get text from textbox 
        user_entered = self.entry.get()
        self.txt.insert(END,"\n"+f"YOU : {user_entered}", 'color')
        self.txt.tag_config("color", background="white", foreground = 'blue')
        
        self.greetings(user_entered)
        




    '''user section '''
    def greetings(self, user_entered):
            
            user_greetings = ['hi', 'hello', 'good morning', 'kaisa hai', 'namaste', "what's up",'hii','hiii','hellooo', 'hey', 'hey!!!', 'holla', 'hola']

            bot_greetings = ['hi', 'hello', 'good morning', 'kaisa hai', 'namaste', "what's up",'hii','hiii','hellooo']

            exit_greet = ['exit', 'bye', 'bye!!!', 'have a good time', 'hmmm','k','ok bye']

            questions_present = ['how many student are present', 'present', 'how many are present', 'present student', 'totall present']

            questions_absent = ['how many student are absent', 'absent', 'how many are absent', 'absent student', 'totall absent']

            highest_present = ['student with highest presents', 'highest presents', 'highest present','least absent']

            highest_absent = ['student with highest absents', 'highest absents', 'highest absent', 'least present']

            date_format = ['totall present today', 'today present', 'yesterday present', 'totall present yesterday', 'present today', 'present yesterday', 'absent today']

            date = ["todays date", "today's date", "date", 'whats todays date']

            about_me = ['who made you', 'what is name of developer']

            about_bot = ['whats your name', 'info', 'who are you', 'how are you', 'details', 'about you']

            
            # continous_present = ['continously present from \d days', 'present from \d days', 'from \d days present']    only for understanding
            format = re.compile(r'(continously)?\s?(present|absent)?\s?from (\d) days\s?(present)?') 
            search = format.search(user_entered.lower())
            #present_specific_date = ['23-01-21', 'info of 23-01-21']    only for understanding
            format1 = re.compile(r'\d\d-\d\d-\d\d')
            search1 = format1.search(user_entered.lower())
            # totall presents of [student's name]
            format2 = re.compile(r'totall? presents? of ([a-zA-Z]+)')
            search2 = format2.search(user_entered.lower())



        
            if user_entered.lower() in user_greetings:

                response = random.choice(bot_greetings)

                self.txt.insert(END,"\n"+f'BOT : {response}')

            elif user_entered.lower() in about_me:

                self.txt.insert(END,"\n"+'BOT : Yash patwa made me using tkinter module in python \nI am glad you asked :P')
                
            
            elif user_entered.lower() in about_bot:

                self.txt.insert(END, "\n"+"BOT : Hello I am Yash's bot\nI am programmed to answer\n20+ Q&A, Excited to help you!!!!")

            elif user_entered.lower() in date:
                date = self.get_todays_date()

                self.txt.insert(END, '\n'+f'BOT : {date}')

            elif user_entered.lower() in date_format:
                format = re.compile(r'(yesterday|today)')
                # print(user_entered.lower())
                search = format.search(user_entered.lower())

                if search.group() == 'yesterday':
                    yesterday_date = get_yesterday_date(search)
                    column_values = self.get_colwith_date(yesterday_date)
                    
                    present, absent = status_attdendance2(column_values)

                    self.txt.insert(END, '\n'+ f'present: {present}, absent: {absent}')

                    lst = ['who made you', 'about you']
                    self.suggestions(lst)
                    
                elif search.group() == 'today':
                     today_date = self.get_todays_date()
                     column_values = self.get_colwith_date(today_date)

                     present, absent = status_attdendance2(column_values)

                     self.txt.insert(END, '\n'+ f'present: {present}, absent: {absent}')

                     lst = ['who made you', 'about you']
                     self.suggestions(lst)


                else:
                    self.txt.insert(END,"\n"+"BOT : sorry, I did not understand!!!")
                    
            
            elif user_entered.lower() in questions_present:
                var1, var2 = self.status_attdendance()

                self.txt.insert(END, "\n"+f'BOT : total {var1} are present')

                lst = ['highest presents', 'highest absents']
                self.suggestions(lst)

                

            elif user_entered.lower() in questions_absent:
                var1, var2 = self.status_attdendance()

                self.txt.insert(END, "\n"+f'BOT : total {var2} are absent')

                lst = ['highest presents', 'highest absents']
                self.suggestions(lst)

                

            elif user_entered.lower() in highest_present:
                list = self.highest_present()
                self.txt.insert(END, "\n"+'BOT : top 3 students with highest present till date are ->')
                for var in list[0:3]:

                    self.txt.insert(END, "\n"+f'{var}')
                lst = ['date', 'todays date']
                self.suggestions(lst)
            
            elif user_entered.lower() in highest_absent:
                list = self.highest_absent()

                self.txt.insert(END, "\n"+'BOT : top 3 students with highest absent till date are ->')
                for var in list[0:3]:

                    self.txt.insert(END, "\n"+f'{var}')

            elif search != None:
                
                if "absent" == str(search.group(2)):
                    name = self.continously_att(int(search.group(3)), 'A')
                    for i in range(int(search.group(3))):

                        self.txt.insert(END, "\n"+"-"+name[i])

                else:
                    name = self.continously_att(int(search.group(3)), 'P')
                    for i in range(int(search.group(3))):

                        self.txt.insert(END, "\n"+"-"+name[i])
            elif search1 != None:
                get_date = search1.group()

                get_date_column_value = self.get_colwith_date(get_date)

                present, absent = status_attdendance2(get_date_column_value)

                self.txt.insert(END, '\n'+ f'present: {present}, absent: {absent}')

                lst = ['totall present yesterday', 'present today']
                self.suggestions(lst)

            elif search2 != None:
                student_name = search2.group(1)
                present, absent = self.excel_data()

                data = present[student_name]
                # print(present)
                self.txt.insert(END, '\n'+f'{student_name} : {data}')

                lst = ['present today', 'present yesterday']
                self.suggestions(lst)


            elif user_entered.lower() in exit_greet:

                response  = random.choice(['holla bye', 'See ya soon', 'Have a great time', 'have a good day', 'bye sir', 'shutting down'])

                self.txt.insert(END, "\n"+f'BOT : {response}')

                
            else:
                self.txt.insert(END,"\n"+"BOT : sorry, I did not understand!!!")

            self.entry.delete(0, END)
                






    ''' questions needs data '''
    #user ask for present student
    def get_todays_date(self):                                      #returns day, you are using bot

        date = self.today.strftime('%d-%m-%y')
        return date

    def status_attdendance(self):                                    #returns totall number of present and absent students
        ws = self.workbook()
        words = self.get_colwith_date(self.get_todays_date())

        total_present = 0
        total_absent = 0

        for row in words:
            if row == "P":
                total_present += 1
            else:
                total_absent += 1

        return total_present, total_absent

    def excel_data(self):                                             #returns dictionary of student name and their corresponding present and absent value
        ws = self.workbook()
        p_value = []
        a_value = []
        # get name of student from excel sheet 
        name = self.get_column(ws, 'A')
        

        for i in range(2,self.row+1):
            p_var = 0
            a_var = 0
            for j in range(2, self.column+1):
                value = ws.cell(column = j , row = i).value
               
                if value == 'P':
                    p_var += 1
                else:
                    a_var += 1
            p_value.append(p_var)
            a_value.append(a_var)
    
        return dict(zip(name, p_value)) ,dict(zip(name, a_value))

    def highest_present(self):                                         #student with highest present
        present_dict, absent_dict = self.excel_data()  

        sorted_dict = sorted(present_dict.items(), key = operator.itemgetter(1), reverse=True)       
        return sorted_dict  
  
    def highest_absent(self):                                          #student with highest absent
        present_dict, absent_dict = self.excel_data()    

        sorted_dict = sorted(absent_dict.items(), key = operator.itemgetter(1), reverse=True)     

        return sorted_dict    

    def get_column(self, ws, alpha):                                   #returns column list with column name as a argument i.e 'A','B'
        
        # print(f'max column {self.column}')
        column = ws[alpha]
        column_list = []
        for i in range(1,self.row):
            column_list.append(column[i].value)
        # print(words)
        return column_list

    def workbook(self):                                                #returns workbook (ws)        
               
            path = os.getcwd() + '\\' + 'Files' + '\\' + self.item_text
            print(path)
            final_file = path + '\\' + 'final.xlsx'
            self.wb = load_workbook(final_file)
            month_name = self.today.strftime('%b')

            ws = self.wb[month_name]
            self.row = ws.max_row
            self.column = ws.max_column
            return ws

    def get_colwith_date(self,str):

            path = os.getcwd() + '\\' + 'Files' + '\\' + self.item_text
            final_file = path + '\\' + 'final.xlsx'

            month_name = self.today.strftime('%b')

            data = pd.read_excel(final_file, engine = 'openpyxl', sheet_name= month_name)
            return data[str]



        
#countinous absent or present from integer times
    def continously_att(self, integer, pa):                           #returns list of particular students
        ws = self.workbook()
        list = []
        
        for i in range(2, self.row+1):
            check_all = []
            j = 0
            for j in range(self.column, self.column-integer, -1):
                # print(j)
                value = ws.cell(column = j, row = i).value
                check_all.append(value)
                # print(value)
            # print(check_all_present)        
            count = 0
            for match in check_all:
                
                if match == pa:
                    count += 1

            if count == integer:
                list.append(ws.cell(column = 1, row = i).value)
        
        return list


    def interface(self):

        root = self.master

        self.txt = Text(root, bd =1, width = 50, height = 10, font=("Arial", 13))
        self.txt.place(x = 0, y = 0,  width = 430, height = 550)

        self.entry = ttk.Entry(root)

        self.entry.bind('<Return>', self.pressed_enter)

        self.entry.place(x = 8, y = 560, width = 340, height = 30 )

        send = ttk.Button(root, text = 'Send', command = self.get_text)
        send.place(x = 350, y = 562)

        scrollbar = Scrollbar(root, command = self.txt.yview)
        scrollbar.place(x =430, y = 0, height = 550)


    def doitall(self):
        
        self.interface()      
        self.send_greeting()
        self.send()

        
        
  
                       
